import openpyxl
book = openpyxl.load_workbook("C:\\Users\\hinan\\Desktop\\Rahul shetty academy python selenium\\PythonDemo1.xlsx")
sheet= book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="Hina"
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)
Dict={}
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="TestCase2":
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i,column=j).value)
            #Dict["Lastname"]=shetty
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)